import math
import os
from dataclasses import dataclass
from datetime import datetime

import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook


# =============================
#  EXCEL (BASE DE DATOS)
# =============================
EXCEL_PATH = "tablas_militares_cijyj_registro.xlsx"
SHEET_NAME = "Inspecciones"

HEADERS = [
    "FechaHora",
    "Operario",
    "Proveedor",
    "Fragancia",
    "NumeroLoteProveedor",
    "LibrasCaneca",
    "NivelInspeccion",
    "Cant_500g", "Codigo_500g", "n_500g",
    "Cant_220g", "Codigo_220g", "n_220g",
    "Cant_30g",  "Codigo_30g",  "n_30g",
    "CalidadEsperada",
    "Observaciones"
]

def ensure_workbook(path=EXCEL_PATH):
    if os.path.exists(path):
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
            wb.save(path)
        else:
            ws = wb[SHEET_NAME]
            if ws.max_row == 0:
                ws.append(HEADERS)
                wb.save(path)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    wb.save(path)

def append_row(row_dict: dict, path=EXCEL_PATH):
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    row = [row_dict.get(h, "") for h in HEADERS]
    ws.append(row)
    wb.save(path)


# -----------------------------
# MIL-STD-105E - TABLE I (letra + n)
# -----------------------------
LOT_RANGES = [
    (2, 8,     {"I": "A", "II": "A", "III": "B"}),
    (9, 15,    {"I": "A", "II": "B", "III": "C"}),
    (16, 25,   {"I": "B", "II": "C", "III": "D"}),
    (26, 50,   {"I": "C", "II": "D", "III": "E"}),
    (51, 90,   {"I": "D", "II": "F", "III": "G"}),
    (91, 150,  {"I": "D", "II": "F", "III": "G"}),
    (151, 280, {"I": "E", "II": "G", "III": "H"}),
    (281, 500, {"I": "F", "II": "H", "III": "J"}),
    (501, 1200, {"I": "G", "II": "J", "III": "K"}),
    (1201, 3200, {"I": "H", "II": "K", "III": "L"}),
    (3201, 10000, {"I": "J", "II": "L", "III": "M"}),
    (10001, 35000, {"I": "K", "II": "M", "III": "N"}),
    (35001, 150000, {"I": "L", "II": "N", "III": "P"}),
    (150001, 500000, {"I": "M", "II": "P", "III": "Q"}),
    (500001, 10**18, {"I": "N", "II": "Q", "III": "R"}),
]

SAMPLE_SIZE_BY_CODE = {
    "A": 2, "B": 3, "C": 5, "D": 8, "E": 13, "F": 20, "G": 32, "H": 50,
    "J": 80, "K": 125, "L": 200, "M": 315, "N": 500, "P": 800, "Q": 1250, "R": 2000
}

# -----------------------------
# AQL informativo (ajustable)
# -----------------------------
DEFAULT_AC_CRIT = 0
DEFAULT_RE_CRIT = 1
DEFAULT_PCT_MAY_AC = 0.03
DEFAULT_PCT_MAY_RE = 0.05
DEFAULT_PCT_MEN_AC = 0.06
DEFAULT_PCT_MEN_RE = 0.10


@dataclass
class PlanMuestreo:
    presentacion: str
    lote: int
    nivel: str
    codigo: str
    n: int
    ac_crit: int
    re_crit: int
    ac_may: int
    re_may: int
    ac_men: int
    re_men: int


def get_code_letter(lot_size: int, level: str) -> str:
    for lo, hi, mapping in LOT_RANGES:
        if lo <= lot_size <= hi:
            return mapping[level]
    return "N"

def get_sample_size(lot_size: int, level: str):
    code = get_code_letter(lot_size, level)
    n = SAMPLE_SIZE_BY_CODE[code]
    return code, min(n, lot_size)

def ac_re_from_pct(n: int, pct_ac: float, pct_re: float):
    ac = math.floor(pct_ac * n)
    re = math.ceil(pct_re * n)
    if re <= ac:
        re = ac + 1
    return ac, re

def build_plan(presentacion: str, lote: int, nivel: str, cfg: dict) -> PlanMuestreo:
    codigo, n = get_sample_size(lote, nivel)
    ac_may, re_may = ac_re_from_pct(n, cfg["pct_may_ac"], cfg["pct_may_re"])
    ac_men, re_men = ac_re_from_pct(n, cfg["pct_men_ac"], cfg["pct_men_re"])
    return PlanMuestreo(
        presentacion=presentacion, lote=lote, nivel=nivel, codigo=codigo, n=n,
        ac_crit=cfg["ac_crit"], re_crit=cfg["re_crit"],
        ac_may=ac_may, re_may=re_may,
        ac_men=ac_men, re_men=re_men
    )

def plan_or_none(presentacion: str, qty: int, nivel: str, cfg: dict):
    if qty <= 0:
        return None
    return build_plan(presentacion, qty, nivel, cfg)

def plans_to_table(p500, p220, p30):
    rows = []
    for label, plan in [("500 g", p500), ("220 g", p220), ("30 g", p30)]:
        if plan is None:
            rows.append({
                "Presentación": label,
                "Lote (und)": 0,
                "Código": "—",
                "Muestra (n)": "—",
                "Crítico (Ac/Re)": "—",
                "Mayor (Ac/Re)": "—",
                "Menor (Ac/Re)": "—",
            })
        else:
            rows.append({
                "Presentación": label,
                "Lote (und)": plan.lote,
                "Código": plan.codigo,
                "Muestra (n)": plan.n,
                "Crítico (Ac/Re)": f"{plan.ac_crit}/{plan.re_crit}",
                "Mayor (Ac/Re)": f"{plan.ac_may}/{plan.re_may}",
                "Menor (Ac/Re)": f"{plan.ac_men}/{plan.re_men}",
            })
    return pd.DataFrame(rows)

def row_fields_from_plan(plan, prefix):
    if plan is None:
        return {f"Cant_{prefix}": 0, f"Codigo_{prefix}": "", f"n_{prefix}": ""}
    return {f"Cant_{prefix}": plan.lote, f"Codigo_{prefix}": plan.codigo, f"n_{prefix}": plan.n}


# =============================
#  STREAMLIT UI (WIZARD)
# =============================
st.set_page_config(page_title="Tablas militares | CI JYJ / Labsens", layout="centered")

# State
if "step" not in st.session_state:
    st.session_state.step = 1
if "plan_df" not in st.session_state:
    st.session_state.plan_df = None
if "plans_obj" not in st.session_state:
    st.session_state.plans_obj = (None, None, None)

# Sidebar config AQL
with st.sidebar:
    st.subheader("⚙️ AQL informativo")
    ac_crit = st.number_input("Crítico - Ac", min_value=0, value=DEFAULT_AC_CRIT, step=1)
    re_crit = st.number_input("Crítico - Re", min_value=1, value=DEFAULT_RE_CRIT, step=1)
    st.markdown("**Mayor (% sobre n)**")
    pct_may_ac = st.number_input("Mayor - % aceptar", min_value=0.0, max_value=1.0, value=DEFAULT_PCT_MAY_AC, step=0.01, format="%.2f")
    pct_may_re = st.number_input("Mayor - % rechazar", min_value=0.0, max_value=1.0, value=DEFAULT_PCT_MAY_RE, step=0.01, format="%.2f")
    st.markdown("**Menor (% sobre n)**")
    pct_men_ac = st.number_input("Menor - % aceptar", min_value=0.0, max_value=1.0, value=DEFAULT_PCT_MEN_AC, step=0.01, format="%.2f")
    pct_men_re = st.number_input("Menor - % rechazar", min_value=0.0, max_value=1.0, value=DEFAULT_PCT_MEN_RE, step=0.01, format="%.2f")

cfg = {
    "ac_crit": int(ac_crit),
    "re_crit": int(re_crit),
    "pct_may_ac": float(pct_may_ac),
    "pct_may_re": float(pct_may_re),
    "pct_men_ac": float(pct_men_ac),
    "pct_men_re": float(pct_men_re),
}

st.title("🧪 Tablas militares (MIL-STD-105E) | CI JYJ / Labsens")
st.caption("Modo asistido por pasos: diligencias → calculas → guardas.")

# Progress
progress_map = {1: 0.33, 2: 0.66, 3: 1.0}
st.progress(progress_map.get(st.session_state.step, 0.33))
st.markdown(f"### Paso {st.session_state.step} de 3")

# Form data storage in session
def init_field(key, default):
    if key not in st.session_state:
        st.session_state[key] = default

init_field("operario", "")
init_field("proveedor", "")
init_field("fragancia", "")
init_field("lote", "")
init_field("libras", 0.0)
init_field("nivel", "II")
init_field("q500", 0)
init_field("q220", 0)
init_field("q30", 0)
init_field("calidad", "Si")
init_field("obs", "")

def next_step():
    st.session_state.step = min(3, st.session_state.step + 1)

def prev_step():
    st.session_state.step = max(1, st.session_state.step - 1)

# STEP 1
if st.session_state.step == 1:
    st.subheader("1) Datos generales")
    st.session_state.operario = st.text_input("Operario (nombre)", value=st.session_state.operario)
    st.session_state.proveedor = st.text_input("Proveedor", value=st.session_state.proveedor)
    st.session_state.fragancia = st.text_input("Fragancia", value=st.session_state.fragancia)
    st.session_state.lote = st.text_input("Número de lote proveedor", value=st.session_state.lote)
    st.session_state.libras = st.number_input("Libras de la caneca (si no aplica, 0)", min_value=0.0, step=0.5, value=float(st.session_state.libras))
    st.session_state.nivel = st.selectbox("Nivel de inspección", ["I", "II", "III"], index=["I","II","III"].index(st.session_state.nivel))

    c1, c2 = st.columns(2)
    with c1:
        st.button("➡️ Siguiente", use_container_width=True, on_click=next_step,
                  disabled=not (st.session_state.operario.strip() and st.session_state.proveedor.strip() and st.session_state.fragancia.strip()))
    with c2:
        st.button("Limpiar", use_container_width=True, on_click=lambda: st.session_state.update({
            "operario":"", "proveedor":"", "fragancia":"", "lote":"", "libras":0.0, "nivel":"II"
        }))

    st.caption("Para continuar: completa Operario, Proveedor y Fragancia.")

# STEP 2
elif st.session_state.step == 2:
    st.subheader("2) Cantidades recibidas")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.session_state.q500 = st.number_input("500 g", min_value=0, step=1, value=int(st.session_state.q500))
    with c2:
        st.session_state.q220 = st.number_input("220 g", min_value=0, step=1, value=int(st.session_state.q220))
    with c3:
        st.session_state.q30 = st.number_input("30 g", min_value=0, step=1, value=int(st.session_state.q30))

    c1, c2 = st.columns(2)
    with c1:
        st.button("⬅️ Atrás", use_container_width=True, on_click=prev_step)
    with c2:
        st.button("➡️ Siguiente", use_container_width=True, on_click=next_step,
                  disabled=(st.session_state.q500 + st.session_state.q220 + st.session_state.q30) == 0)

    st.caption("Para continuar: al menos una cantidad debe ser > 0.")

# STEP 3
else:
    st.subheader("3) Calcular, revisar y guardar")

    # Resumen rápido arriba (se siente como “web app”)
    with st.container(border=True):
        st.markdown("**Resumen de recepción**")
        st.write(f"👤 Operario: **{st.session_state.operario}**")
        st.write(f"🏷️ Proveedor: **{st.session_state.proveedor}**")
        st.write(f"🌸 Fragancia: **{st.session_state.fragancia}**")
        st.write(f"🔢 Lote proveedor: **{st.session_state.lote or 'N/A'}**")
        st.write(f"⚖️ Libras caneca: **{st.session_state.libras if st.session_state.libras else 'N/A'}**")
        st.write(f"📏 Nivel inspección: **{st.session_state.nivel}**")

    st.markdown("**Cantidades**")
    st.write(f"500 g: **{st.session_state.q500}** | 220 g: **{st.session_state.q220}** | 30 g: **{st.session_state.q30}**")

    st.session_state.calidad = st.selectbox("¿La fragancia cumplió con la calidad esperada?", ["Si", "No"], index=["Si","No"].index(st.session_state.calidad))
    st.session_state.obs = st.text_area("Observaciones", value=st.session_state.obs, placeholder="Ej: N/A o detalle de hallazgos")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.button("⬅️ Atrás", use_container_width=True, on_click=prev_step)
    with c2:
        if st.button("📌 Calcular plan", use_container_width=True):
            p500 = plan_or_none("500 g", int(st.session_state.q500), st.session_state.nivel, cfg)
            p220 = plan_or_none("220 g", int(st.session_state.q220), st.session_state.nivel, cfg)
            p30  = plan_or_none("30 g",  int(st.session_state.q30),  st.session_state.nivel, cfg)
            df = plans_to_table(p500, p220, p30)
            st.session_state.plan_df = df
            st.session_state.plans_obj = (p500, p220, p30)

    with c3:
        if st.button("💾 Guardar registro", use_container_width=True, disabled=st.session_state.plan_df is None):
            p500, p220, p30 = st.session_state.plans_obj
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            row = {
                "FechaHora": now,
                "Operario": st.session_state.operario.strip(),
                "Proveedor": st.session_state.proveedor.strip(),
                "Fragancia": st.session_state.fragancia.strip(),
                "NumeroLoteProveedor": (st.session_state.lote or "").strip(),
                "LibrasCaneca": st.session_state.libras if st.session_state.libras != 0 else "",
                "NivelInspeccion": st.session_state.nivel,
                "CalidadEsperada": st.session_state.calidad,
                "Observaciones": st.session_state.obs.strip() if st.session_state.obs else ""
            }

            row.update(row_fields_from_plan(p500, "500g"))
            row.update(row_fields_from_plan(p220, "220g"))
            row.update(row_fields_from_plan(p30,  "30g"))

            append_row(row)
            st.success(f"✅ Registro guardado (1 fila). Archivo: {EXCEL_PATH}")

            if os.path.exists(EXCEL_PATH):
                with open(EXCEL_PATH, "rb") as f:
                    st.download_button(
                        "⬇️ Descargar Excel actualizado",
                        data=f,
                        file_name=EXCEL_PATH,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

    if st.session_state.plan_df is not None:
        st.markdown("---")
        st.subheader("📌 Plan de muestreo (código, muestra n y AQL informativo)")
        st.dataframe(st.session_state.plan_df, use_container_width=True, hide_index=True)
        st.caption("Crítico/Major/Menor se muestran como Ac/Re (aceptación / rechazo).")
        st.caption("Tip: si quieres que se vea más “dashboard”, lo pasamos a tarjetas por presentación.")

